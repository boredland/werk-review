import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DATA_DIR = Path(__file__).parent.parent / "data"
OUT = Path(__file__).parent / "example-sheets" / "werk-review.xlsx"

wb = Workbook()

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def style_header(ws):
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def sources_to_str(sources):
    if not sources:
        return ""
    return "|".join(f'{s["label"]} <{s["url"]}>' for s in sources)


# --- Autoren ---
ws = wb.active
ws.title = "Autoren"
headers = ["id", "name", "slug", "aliases", "born", "died", "bio", "source_note"]
ws.append(headers)

for f in sorted(DATA_DIR.glob("authors/*.json")):
    a = json.loads(f.read_text())
    ws.append([
        a["id"], a["name"], a["slug"],
        "|".join(a.get("aliases", [])),
        a.get("born"), a.get("died"),
        a.get("bio", ""),
        a.get("source_note", ""),
    ])

style_header(ws)
ws.column_dimensions["B"].width = 25
ws.column_dimensions["G"].width = 60

# --- Werke ---
ws = wb.create_sheet("Werke")
headers = ["id", "author_id", "genre_ids", "title", "slug", "aliases", "year_display",
           "parent_slugs", "fortsetzung_von_ids", "plot", "source_note", "sources"]
ws.append(headers)

works = []
for f in sorted(DATA_DIR.glob("works/*.json")):
    works.append(json.loads(f.read_text()))
works.sort(key=lambda w: w.get("year") or 9999)

for w in works:
    ws.append([
        w["id"], w["author_id"], "|".join(w.get("genre_ids", [])),
        w["title"], w["slug"], "|".join(w.get("aliases", [])),
        w.get("year_display", ""),
        "|".join(w.get("parent_slugs", [])),
        "|".join(w.get("fortsetzung_von_ids", [])),
        w.get("plot"),
        w.get("source_note", ""),
        sources_to_str(w.get("sources", [])),
    ])

style_header(ws)
ws.column_dimensions["D"].width = 40
ws.column_dimensions["J"].width = 50

# --- Genres ---
ws = wb.create_sheet("Genres")
headers = ["id", "name", "slug"]
ws.append(headers)

genres = json.loads((DATA_DIR / "genres.json").read_text())
for g in genres:
    ws.append([g["id"], g["name"], g["slug"]])

style_header(ws)
ws.column_dimensions["B"].width = 20

# --- Links (enriched external links, one row per link; work_slug is the filename) ---
ws = wb.create_sheet("Links")
headers = ["work_slug", "source", "format", "url", "label"]
ws.append(headers)

link_count = 0
for f in sorted(DATA_DIR.glob("links/*.json")):
    links = json.loads(f.read_text())
    if not isinstance(links, list):
        continue
    for l in links:
        ws.append([f.stem, l.get("source"), l.get("format"), l.get("url"), l.get("label")])
        link_count += 1

style_header(ws)
ws.column_dimensions["A"].width = 35
ws.column_dimensions["D"].width = 60
ws.column_dimensions["E"].width = 40

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Created {OUT}")
print(f"  Autoren: {wb['Autoren'].max_row - 1} rows")
print(f"  Werke: {wb['Werke'].max_row - 1} rows")
print(f"  Genres: {wb['Genres'].max_row - 1} rows")
print(f"  Links: {link_count} rows")
